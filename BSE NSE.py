import pandas as pd
from openpyxl import workbook
from openpyxl import load_workbook

bsewb = load_workbook (r'C:\\Users\\admin\Desktop\\NSE\StockAnalyzer-main\\BSE.xlsx')
nsewb = load_workbook (r'C:\\Users\\admin\Desktop\\NSE\StockAnalyzer-main\\NSE.xlsx')

bsews = bsewb['Equity']
nsews = nsewb['NSE']
bsedata = bsews.values
nsedata = nsews.values
bsecolumn = next(bsedata)[0:]
nsecolumn = next(nsedata)[0:]
bsedf = pd.DataFrame(bsedata,columns = bsecolumn)
nsedf = pd.DataFrame(nsedata,columns = nsecolumn)

nsenewws = bsewb.create_sheet('nsenew')
nsenewws.title = 'nsenew'
nsenewws.append(nsecolumn)
for secid in nsedf['Security Id'].unique():
    filterdf = bsedf.query('`Security Id` == @secid')
    if not filterdf.empty:
        finalfilterdf = nsedf.query('`Security Id` == @secid')
        #print([finalfilterdf['ISIN No'].values[0],finalfilterdf['Security Id'].values[0],finalfilterdf['Security Name'].values[0],filterdf['Industry'],filterdf['Group']])
        print(filterdf[['Industry', 'Group']])

