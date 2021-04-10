import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

# let's read the sectors data from workbook
Sectorswb = load_workbook(r'D:\NSEData\Sectors.xlsx')

# let's read the Company Symbol data from workbook
CompanySymbolwb = load_workbook(r'D:\NSEData\CompanySymbol.xlsx')

# let's output the sheet names
#print(Sectorswb.sheetnames)

# let's grab the sheet from which we need to read the values
ws = Sectorswb['Sheet2']

# let's grab the sheet from which we need to read the company symbols
CompanySymbolws = CompanySymbolwb['in']

# let's assign the whole sheet to data for building of data frame
data = ws.values

# let's assign the whole sheet to data for building of data frame for Company Symbols
CompanySymboldata = CompanySymbolws.values

# let's grab the header/column values for building data frame
columns = next(data)[0:]

# let's grab the header/column values for building data frame for Company Symbols
CompanySymbolcolumns = next(CompanySymboldata)[0:]


# let's build the data frame from data and columns extracted above
df = pd.DataFrame(data, columns=columns)

# let's build the data frame from data and columns for above extracted Company Symbols
CompanySymboldf = pd.DataFrame(CompanySymboldata, columns=CompanySymbolcolumns)

# let's print all the unique sectors
#print(df['Sector'].unique())

# create new work book
workbook = Workbook()

headers = ['Company','Symbol']
CompaniesMapped = 0
CompaniesUnMapped = 0

# let's iterate the unique sectors
for sector in df['Sector'].unique():
    print(sector)
    # check if the sector is not null or empty
    if sector is not None:
        worksheet = workbook.create_sheet(sector)
        worksheet.title = sector
        worksheet.append(headers)
        currentSector = sector
        filteredDf = df.query("Sector == @currentSector")
        # let's iterate the companies belonging to currently iterated sector
        for company in filteredDf['Company Name'].unique():
            
            # let's try to first match the whole company name
            CompanySymbolfilteredDf = CompanySymboldf.query("Company == @company")
            
            # still if our dataframe is empty - as the company name didn't the match
            # Hence, now try with appending the "Limited" keyword
            if CompanySymbolfilteredDf.empty:
                companylimited = str(company) + ' Limited'
                CompanySymbolfilteredDf = CompanySymboldf.query("Company == @companylimited")

            # still if our dataframe is empty - as the company name didn't the match
            # Hence, now try with appending the " (I) Limited" keyword
            if CompanySymbolfilteredDf.empty:
                companyIlimited = str(company) + ' (I) Limited'
                CompanySymbolfilteredDf = CompanySymboldf.query("Company == @companyIlimited")

            # still if our dataframe is empty - as the company name didn't the match
            # Hence, now try with contains method which is available as a part of string comparision
            if CompanySymbolfilteredDf.empty:
                CompanySymbolfilteredDf = CompanySymboldf[CompanySymboldf['Company'].str.contains(company)]

            # Now if we don't have empty dataframe - which implies that we found the company's symbol 
            if not CompanySymbolfilteredDf.empty:
                CompaniesMapped = CompaniesMapped + 1
                worksheet.append([company, CompanySymbolfilteredDf['Symbol'].values[0]])
            else:
                CompaniesUnMapped = CompaniesUnMapped + 1
                worksheet.append([company, ''])

print("Total Companies processed: " + str(CompaniesMapped + CompaniesUnMapped))
print("Total Companies code successfully found: " + str(CompaniesMapped))
print("Total Companies code successfully not found: " + str(CompaniesUnMapped))
# Grab the currently active sheet
#sheet = workbook.active

# let's setup the path to save the Sector wise companies
saveSectorWisefilepath = r'D:\NSEData\SectorWiseCompanies.xlsx'

#df.to_excel(saveSectorWisefilepath)

# Let's save the workbook
workbook.save(saveSectorWisefilepath)

#df = pd.read_excel(r'D:\NSEData\Sectors.xlsx')
#print (df['Sector '].unique())