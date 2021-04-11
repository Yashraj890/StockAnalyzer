import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

# let's read the sectors data from workbook
Sectorswb = load_workbook(r'D:\NSEData\Sectors.xlsx')

# let's read the sectorswise company data from workbook
SectorwiseCompanieswb = load_workbook(r'D:\NSEData\SectorWiseCompanies.xlsx')

# let's read the sec bhav data full company data from workbook
#secBhavDatawb = load_workbook(r'D:\NSEData\sec_bhavdata_full.csv')

# let's grab the sheet from which we need to read the values
ws = Sectorswb['Sheet2']

# let's grab the sheet from which we need to daily records
#secBhavDataws = secBhavDatawb['sec_bhavdata_full']

# let's assign the whole sheet to data for building of data frame
data = ws.values

# let's assign the whole sheet to data for building of data frame
#secBhavData = secBhavDataws.values

# let's grab the header/column values for building data frame
columns = next(data)[0:]

# let's grab the header/column values for building data frame
secBhavDatacolumns = next(data)[0:]

# let's build the data frame from data and columns extracted above
df = pd.DataFrame(data, columns=columns)

#secBhavDatadf = pd.DataFrame(secBhavData, columns=secBhavDatacolumns)
secBhavDatadf = pd.read_csv(r'D:\NSEData\sec_bhavdata_full.csv')

for sector in df['Sector'].unique():
    if sector is not None:
        currentSector = sector
        sectorFilePath = r'D:\\NSEData\\' + currentSector + '.xlsx'
        # let's read the sectors data from workbook
        CurrentSectorwb = load_workbook(sectorFilePath)
        for currCompanySymbol in CurrentSectorwb.sheetnames:
            currSymbol = currCompanySymbol
            seriesFilter = ' EQ'
            filteredDf = secBhavDatadf.query('SYMBOL == @currSymbol & SERIES == @seriesFilter')
            #print(filteredDf)
            if not filteredDf.empty:
                    with pd.ExcelWriter(sectorFilePath, engine='openpyxl', mode='a') as writer: # pylint: disable=abstract-class-instantiated
                        writer.book = CurrentSectorwb
                        writer.sheets = dict((ws.title, ws) for ws in CurrentSectorwb.worksheets)
                        filteredDf.to_excel(writer, sheet_name = currSymbol)
                        writer.save()
                        #writer.close()
                    



