import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

# let's read the sectors data from workbook
Sectorswb = load_workbook(r'D:\NSEData\Sectors.xlsx')

# let's read the sectorswise company data from workbook
SectorwiseCompanieswb = load_workbook(r'D:\NSEData\SectorWiseCompanies.xlsx')

# let's grab the sheet from which we need to read the values
ws = Sectorswb['Sheet2']

# let's assign the whole sheet to data for building of data frame
data = ws.values

# let's grab the header/column values for building data frame
columns = next(data)[0:]

# let's build the data frame from data and columns extracted above
df = pd.DataFrame(data, columns=columns)

headers = ['Company','Symbol','DATE1','PREV_CLOSE','OPEN_PRICE','HIGH_PRICE','LOW_PRICE','LAST_PRICE','CLOSE_PRICE','AVG_PRICE','TTL_TRD_QNTY','TURNOVER_LACS','NO_OF_TRADES','DELIV_QTY','DELIV_PER']

# let's iterate the unique sectors
for sector in df['Sector'].unique():
    if sector is not None:
        currSector = ""
        currentSector = sector

        # create new work book for current sector
        workbook = Workbook()

        sheetIndex = 0
        # let's loop all the sector wise company sheets to find the 
        # one matching with our current iterating sector and set it active
        for sheetname in SectorwiseCompanieswb.sheetnames:
            currSector = sector[0:31]
            if (currSector.lower() == sheetname.lower()):
                SectorwiseCompanieswb.active = sheetIndex
                break
            sheetIndex = sheetIndex + 1
        
        # let's grab the current sector sheet from which we need to read the values
        SectorwiseCompaniesws = SectorwiseCompanieswb.active
        
        # let's assign the whole sheet to data for building of data frame
        SectorwiseCompaniesdata = SectorwiseCompaniesws.values

        # let's grab the header/column values for building data frame
        SectorwiseCompaniescolumns = next(SectorwiseCompaniesdata)[0:]

        # let's build the data frame from data and columns extracted above
        SectorwiseCompaniesdf = pd.DataFrame(SectorwiseCompaniesdata, columns=SectorwiseCompaniescolumns)
        
        # let's iterate the unique symbols
        for symbol in SectorwiseCompaniesdf['Symbol'].unique():
            if symbol is not None:
                print(symbol)
                worksheet = workbook.create_sheet(symbol)
                worksheet.title = symbol
                worksheet.append(headers)
    # let's setup the file path where we have saved the Sector wise companies
    SectorWisefilepath = r'D:\\NSEData\\'
    SectorWisefilepath = SectorWisefilepath + currentSector + '.xlsx'

    # let's remove the default sheet from the workbook before saving
    workbook.remove_sheet('sheet1')

    # Let's save the workbook
    workbook.save(SectorWisefilepath)
